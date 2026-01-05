import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import re
import datetime
import os

# --- Configuration ---
# Using the filename found in the workspace
INPUT_CSV = r"C:\Users\DKS\Desktop\VSC\Dec25\python\Course Schedule 2019-202 Semester(Pre-Registration).csv"
OUTPUT_XLSX = r"C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Dashboard_v2.xlsx"

# --- Constants ---
DAYS_MAP = {
    'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri', 'S': 'Sat',
    'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat'
}
DAY_ORDER = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

# --- Parsing Logic ---

def parse_day_string(day_str):
    """
    Parses a string like 'MThF', 'MW', 'T', 'Th' into a list of standard day names.
    """
    if not day_str:
        return []
    
    # Regex to match 'Th' or single letters M, T, W, F, S
    # We must match 'Th' first before 'T'
    tokens = re.findall(r'Th|M|T|W|F|S', day_str)
    days = []
    for t in tokens:
        if t in DAYS_MAP:
            days.append(DAYS_MAP[t])
    return days

def parse_time_field(time_val, default_venue):
    """
    Parses a complex time string from the CSV.
    Returns a list of dictionaries:
    [{'day': 'Mon', 'start': '09:00', 'end': '10:00', 'venue': 'L11'}, ...]
    """
    if pd.isna(time_val) or str(time_val).strip() == "":
        return []

    raw_str = str(time_val).strip()
    meetings = []
    
    # Split by comma if it looks like multiple distinct definitions
    # e.g. "T 09:00-10:15 ,M 12:00-13:15"
    # But be careful not to split inside parens if that ever happens (unlikely for this data)
    segments = raw_str.split(',')
    
    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue
            
        # 1. Extract Time Range (HH:MM-HH:MM)
        # Look for pattern like 09:00-10:00 or 9:00-10:00
        time_match = re.search(r'(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})', segment)
        
        if not time_match:
            # Log error or skip? We will skip but maybe return a "parse error" flag if needed
            continue
            
        start_time = time_match.group(1)
        end_time = time_match.group(2)
        
        # Normalize times to HH:MM (add leading zero if needed)
        def normalize_time(t):
            h, m = t.split(':')
            return f"{int(h):02d}:{m}"
        
        start_time = normalize_time(start_time)
        end_time = normalize_time(end_time)
        
        # 2. Remove the time part to process days/venues
        remainder = segment.replace(time_match.group(0), "").strip()
        
        # 3. Strategy: Look for "Day (Venue)" patterns first
        # Regex: ([M|T|W|Th|F|S]+)\s*\(([^)]+)\)
        # This matches things like "M (L11)" or "MW (L7)"
        
        # We iterate through the string finding these patterns
        # If we find them, we consume them.
        # If text remains that is just days without parens, we use default_venue.
        
        # Pattern for DayToken + Optional Venue in Parens
        # We need to be careful. "M (L11) W (L11)" -> M@L11, W@L11
        # "MW" -> M@Default, W@Default
        
        # Let's try to tokenize the remainder
        # We can split by spaces, but parens might contain spaces.
        
        # Alternative: Find all (DayChunk, VenueChunk) pairs
        # Regex to find "Days (Venue)"
        # Note: Days can be "M", "Th", "MW", "MTh"
        day_venue_pattern = re.compile(r'((?:Th|M|T|W|F|S)+)\s*\(([^)]+)\)')
        
        matches = list(day_venue_pattern.finditer(remainder))
        
        matched_spans = []
        
        # Process explicit day-venue pairs
        for m in matches:
            d_str = m.group(1)
            v_str = m.group(2)
            matched_spans.append(m.span())
            
            days = parse_day_string(d_str)
            for d in days:
                meetings.append({
                    'day': d,
                    'start': start_time,
                    'end': end_time,
                    'venue': v_str.strip()
                })
        
        # Process remaining parts (days without explicit venue)
        # We remove the matched parts from remainder
        # Construct a mask or just replace with spaces?
        temp_remainder = list(remainder)
        for start, end in matched_spans:
            for i in range(start, end):
                temp_remainder[i] = ' '
        
        clean_remainder = "".join(temp_remainder).strip()
        
        # Now parse days from the clean remainder
        if clean_remainder:
            days = parse_day_string(clean_remainder)
            venue_to_use = default_venue if pd.notna(default_venue) else "TBA"
            for d in days:
                meetings.append({
                    'day': d,
                    'start': start_time,
                    'end': end_time,
                    'venue': str(venue_to_use).strip()
                })
                
    return meetings

# --- Main Processing ---

def main():
    print("Step A: Reading and Normalizing Data...")
    
    try:
        df = pd.read_csv(INPUT_CSV)
    except FileNotFoundError:
        print(f"Error: File not found at {INPUT_CSV}")
        return

    meetings_rows = []
    parse_errors = []
    
    # Identify Time/Venue columns
    # The CSV likely has "Time", "Venue", "Time.1", "Venue.1", etc.
    time_cols = [c for c in df.columns if str(c).startswith('Time')]
    
    for idx, row in df.iterrows():
        # 1. Basic Course Info
        raw_name = str(row.get('Course Name/Group Name', ''))
        instructor = str(row.get('Instructor', ''))
        
        # Extract Code and Title
        # Format: "TITLE(CODE)" or just "TITLE"
        match = re.search(r'(.*?)\((.*?)\)$', raw_name)
        if match:
            course_title = match.group(1).strip()
            course_code = match.group(2).strip()
        else:
            course_title = raw_name.strip()
            course_code = raw_name.strip() # Fallback
            
        display_course = f"{course_code} - {course_title}"
        
        # 2. Process Time/Venue Pairs
        has_meetings = False
        
        for t_col in time_cols:
            # Find matching venue column
            suffix = t_col.replace('Time', '')
            v_col = 'Venue' + suffix
            
            time_val = row.get(t_col)
            venue_val = row.get(v_col)
            
            if pd.isna(time_val) or str(time_val).strip() == "":
                continue
                
            parsed = parse_time_field(time_val, venue_val)
            
            if not parsed:
                # Log error if time existed but parsing failed
                parse_errors.append({
                    'Row': idx + 2,
                    'Course': display_course,
                    'RawTime': time_val,
                    'RawVenue': venue_val
                })
            else:
                has_meetings = True
                for m in parsed:
                    meetings_rows.append({
                        'CourseCode': course_code,
                        'CourseTitle': course_title,
                        'DisplayCourse': display_course,
                        'Instructor': instructor,
                        'Day': m['day'],
                        'StartTime': m['start'],
                        'EndTime': m['end'],
                        'Venue': m['venue']
                    })
        
        if not has_meetings:
            # Add a placeholder for "No Scheduled Slot" logic if desired
            # For now, we just ensure the course exists in the course list
            pass

    # Create DataFrames
    meetings_df = pd.DataFrame(meetings_rows)
    
    # Create Unique Course List
    # We need to ensure ALL courses are in the course list, even if they have no meetings
    # So we extract unique courses from the original DF, not just meetings_df
    course_list_data = []
    seen_codes = set()
    
    for _, row in df.iterrows():
        raw_name = str(row.get('Course Name/Group Name', ''))
        match = re.search(r'(.*?)\((.*?)\)$', raw_name)
        if match:
            title = match.group(1).strip()
            code = match.group(2).strip()
        else:
            title = raw_name.strip()
            code = raw_name.strip()
            
        if code not in seen_codes:
            course_list_data.append({
                'CourseCode': code,
                'CourseTitle': title,
                'DisplayCourse': f"{code} - {title}",
                'Instructor': row.get('Instructor', '')
            })
            seen_codes.add(code)
            
    courses_df = pd.DataFrame(course_list_data)
    courses_df.sort_values('DisplayCourse', inplace=True)
    
    print(f"Extracted {len(courses_df)} unique courses.")
    print(f"Generated {len(meetings_df)} meeting rows.")
    if parse_errors:
        print(f"Encountered {len(parse_errors)} parsing issues (logged to Excel).")

    # --- Step B: Create Excel Workbook ---
    print("Step B: Building Excel Workbook...")
    wb = Workbook()
    
    # 1. MasterCourses Sheet
    ws_courses = wb.active
    ws_courses.title = "MasterCourses"
    
    # Write Headers & Data
    ws_courses.append(list(courses_df.columns))
    for r in dataframe_to_rows(courses_df, index=False, header=False):
        ws_courses.append(r)
        
    # Create Table
    create_table(ws_courses, "tblCourses", len(courses_df))
    
    # 2. MasterMeetings Sheet
    ws_meetings = wb.create_sheet("MasterMeetings")
    ws_meetings.append(list(meetings_df.columns))
    for r in dataframe_to_rows(meetings_df, index=False, header=False):
        ws_meetings.append(r)
    create_table(ws_meetings, "tblMeetings", len(meetings_df))
    
    # 3. ParseErrors Sheet
    if parse_errors:
        ws_errors = wb.create_sheet("ParseErrors")
        ws_errors.append(['Row', 'Course', 'RawTime', 'RawVenue'])
        for err in parse_errors:
            ws_errors.append([err['Row'], err['Course'], err['RawTime'], err['RawVenue']])

    # --- Step C: Dashboard UI ---
    print("Step C: Designing Dashboard...")
    ws_dash = wb.create_sheet("Dashboard")
    wb.move_sheet(ws_dash, offset=-10) # Move to first position
    ws_dash.sheet_view.showGridLines = False
    
    # Styles
    NAVY_FILL = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    BORDER = Border(left=Side(style='thin', color="D9D9D9"), 
                    right=Side(style='thin', color="D9D9D9"), 
                    top=Side(style='thin', color="D9D9D9"), 
                    bottom=Side(style='thin', color="D9D9D9"))
    
    # Title
    ws_dash['A1'] = "IITK Course Timetable Dashboard"
    ws_dash['A1'].font = Font(size=20, bold=True, color="203764")
    
    # --- Search & Selection Section ---
    ws_dash['B3'] = "Search Course:"
    ws_dash['B3'].font = BOLD_FONT
    ws_dash['B3'].alignment = Alignment(horizontal='right')
    
    # Search Box (C3 is actually part of the selection area in the prompt, 
    # but let's put the search box in B4 and selections in C4:C9 to match prompt logic roughly)
    # Prompt says: "Search box in B3", "Selectors C3:C8". Okay.
    
    # Search Input
    ws_dash['B4'] = "Type to Search:"
    ws_dash['B4'].font = Font(italic=True, size=9)
    ws_dash['B4'].alignment = Alignment(horizontal='right', vertical='top')
    
    # We'll use B3 as the input cell for search text
    ws_dash['C3'].value = "" # Placeholder for user typing
    # Actually, prompt says "Search box in B3". Let's make B3 the input.
    ws_dash['B3'].value = "Search Filter:"
    ws_dash['C3'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_dash['C3'].border = BORDER
    
    # Selectors C5:C10 (Moving down slightly to make room)
    ws_dash['B5'] = "Select Courses (Max 6):"
    ws_dash['B5'].font = BOLD_FONT
    
    selector_start_row = 5
    for i in range(6):
        r = selector_start_row + i
        cell = ws_dash.cell(row=r, column=3) # C
        cell.border = BORDER
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Instructor Info
        ws_dash.cell(row=r, column=4).value = "Instructor:"
        ws_dash.cell(row=r, column=4).font = Font(size=8, color="808080")
        ws_dash.cell(row=r, column=4).alignment = Alignment(horizontal='right')
        
        inst_cell = ws_dash.cell(row=r, column=5) # E
        # Formula to fetch instructor
        # =IFERROR(VLOOKUP(C5, tblCourses[[DisplayCourse]:[Instructor]], 4, FALSE), "")
        # Using XLOOKUP for robustness
        inst_cell.value = f'=IF(C{r}="","",XLOOKUP(C{r},tblCourses[DisplayCourse],tblCourses[Instructor],""))'
        inst_cell.font = Font(italic=True)

    # --- Search Logic (Spilled Array) ---
    # We'll use a hidden area or far right column for the filtered list
    # Let's use Column AA on MasterCourses to keep Dashboard clean
    # Formula in MasterCourses!AA1:
    # =FILTER(tblCourses[DisplayCourse], ISNUMBER(SEARCH(Dashboard!C3, tblCourses[DisplayCourse])), tblCourses[DisplayCourse])
    # If C3 is empty, it returns everything.
    
    ws_courses['AA1'].value = "Filtered List"
    ws_courses['AA2'].value = '=SORT(FILTER(tblCourses[DisplayCourse], ISNUMBER(SEARCH(Dashboard!$C$3, tblCourses[DisplayCourse])), tblCourses[DisplayCourse]))'
    
    # Data Validation for Selectors
    # Reference the spilled range MasterCourses!$AA$2#
    dv = DataValidation(type="list", formula1='=MasterCourses!$AA$2#', allow_blank=True)
    ws_dash.add_data_validation(dv)
    for i in range(6):
        dv.add(ws_dash.cell(row=selector_start_row+i, column=3))

    # --- Weekly Grid ---
    grid_header_row = 12
    grid_cols = ['B', 'C', 'D', 'E', 'F'] # Mon-Fri
    
    # Headers
    for i, day in enumerate(DAY_ORDER[:5]): # Mon-Fri
        cell = ws_dash.cell(row=grid_header_row, column=2+i)
        cell.value = day
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    # Time Slots (15 min intervals)
    start_time = datetime.datetime(2000, 1, 1, 8, 0) # 08:00
    end_time = datetime.datetime(2000, 1, 1, 19, 30) # 19:30
    
    current_time = start_time
    row_idx = grid_header_row + 1
    
    while current_time <= end_time:
        time_str = current_time.strftime("%H:%M")
        
        # Time Label Column A
        cell_a = ws_dash.cell(row=row_idx, column=1)
        cell_a.value = time_str
        cell_a.font = Font(size=9)
        cell_a.fill = GRAY_FILL
        cell_a.alignment = Alignment(horizontal='right')
        cell_a.border = BORDER
        
        # Grid Cells
        for col_idx in range(2, 7): # B-F
            cell = ws_dash.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
            
            # --- THE FORMULA ---
            # We need to inject the complex formula here.
            # To make it readable and robust, we use LET.
            
            col_letter = get_column_letter(col_idx)
            day_ref = f"{col_letter}${grid_header_row}" # e.g. B$12
            time_ref = f"$A{row_idx}" # e.g. $A13
            
            # Formula Logic:
            # 1. Define SlotStart and SlotEnd (Time values)
            #    Note: Excel sees "08:00" string as text unless converted. 
            #    But our A column is text "08:00". We can use TIMEVALUE.
            # 2. Get Selected Courses (C5:C10)
            # 3. Filter Meetings
            
            # We need to handle the "CourseCode" extraction from "DisplayCourse" inside the formula
            # DisplayCourse format: "CODE - TITLE"
            # Code = LEFT(..., FIND(" - ")-1)
            
            # Range for selectors: $C$5:$C$10
            
            formula = (
                f'=LET('
                f'SlotStart, TIMEVALUE({time_ref}), '
                f'SlotEnd, SlotStart + TIME(0,15,0), '
                f'DayHeader, {day_ref}, '
                f'SelectedList, FILTER($C$5:$C$10, $C$5:$C$10<>""), '
                f'SelectedCodes, IF(ISERROR(SelectedList), "", LEFT(SelectedList, FIND(" - ", SelectedList&" - ")-1)), '
                f'Meetings, FILTER(tblMeetings, (tblMeetings[Day]=DayHeader) * (ISNUMBER(XMATCH(tblMeetings[CourseCode], SelectedCodes))) * (TIMEVALUE(tblMeetings[StartTime])<SlotEnd) * (TIMEVALUE(tblMeetings[EndTime])>SlotStart), ""), '
                f'Result, IF(ISERROR(INDEX(Meetings,1,1)), "", TEXTJOIN(", ", TRUE, UNIQUE(INDEX(Meetings,,1) & CHAR(10) & "@ " & INDEX(Meetings,,8)))), '
                f'IF(ISNUMBER(SEARCH(",", Result)), "CONFLICT: " & Result, Result)'
                f')'
            )
            
            cell.value = formula
            
        current_time += datetime.timedelta(minutes=15)
        row_idx += 1

    # --- Conditional Formatting ---
    # Grid Range
    grid_range = f"B{grid_header_row+1}:F{row_idx-1}"
    
    # 1. Conflict (Red)
    ws_dash.conditional_formatting.add(grid_range, 
        FormulaRule(formula=[f'ISNUMBER(SEARCH("CONFLICT", B{grid_header_row+1}))'], stopIfTrue=True, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    # 2. Per-Course Colors (6 colors)
    colors = ["DDEBF7", "E2EFDA", "FFF2CC", "FCE4D6", "E7E6E6", "D9E1F2"] # Blue, Green, Yellow, Orange, Grey, Purple-ish
    
    for i in range(6):
        sel_row = selector_start_row + i
        color = colors[i]
        # Formula: Check if the cell contains the Course Code from selector i
        # Code = LEFT($C$5, FIND(" - ")-1)
        # We check if that code is inside the cell value
        
        rule_formula = f'AND($C${sel_row}<>"", ISNUMBER(SEARCH(LEFT($C${sel_row}, FIND(" - ", $C${sel_row}&" - ")-1), B{grid_header_row+1})))'
        
        ws_dash.conditional_formatting.add(grid_range,
            FormulaRule(formula=[rule_formula], stopIfTrue=True, fill=PatternFill(start_color=color, end_color=color, fill_type="solid")))

    # --- Final Touches ---
    ws_dash.column_dimensions['A'].width = 10
    for c in ['B', 'C', 'D', 'E', 'F']:
        ws_dash.column_dimensions[c].width = 25
        
    ws_dash.freeze_panes = f"B{grid_header_row+1}"
    
    # Save
    wb.save(OUTPUT_XLSX)
    print(f"Success! Dashboard v2 saved to: {OUTPUT_XLSX}")

def dataframe_to_rows(df, index=True, header=True):
    if header:
        yield [index_name for index_name in df.index.names] + list(df.columns) if index else list(df.columns)
    for row in df.itertuples(index=index, name=None):
        yield [None if pd.isna(x) else x for x in row]

def create_table(ws, table_name, num_rows):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    if num_rows == 0: return
    
    max_col = get_column_letter(ws.max_column)
    ref = f"A1:{max_col}{num_rows+1}"
    
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

if __name__ == "__main__":
    main()
