import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import datetime
import os

# --- Configuration ---
# Using the filename found in the workspace
INPUT_CSV = r"C:\Users\DKS\Desktop\VSC\Dec25\python\Course Schedule 2019-202 Semester(Pre-Registration).csv"
OUTPUT_XLSX = r"C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Dashboard_v3.xlsx"

# --- Constants ---
DAYS_MAP = {
    'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri', 'S': 'Sat',
    'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat'
}
DAY_ORDER = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

# --- Parsing Logic ---

def parse_day_string(day_str):
    """
    Parses a string like 'MThF', 'MW', 'T', 'Th' into a list of standard day names.
    """
    if not day_str:
        return []
    
    # Regex to match 'Th' or single letters M, T, W, F, S
    tokens = re.findall(r'Th|M|T|W|F|S', str(day_str))
    days = []
    for t in tokens:
        if t in DAYS_MAP:
            days.append(DAYS_MAP[t])
    return days

def parse_time_field(time_val, default_venue):
    """
    Parses a complex time string from the CSV.
    Returns a list of dictionaries:
    [{'day': 'Mon', 'start': '09:00', 'end': '10:00', 'venue': 'L11'}, ...]
    """
    if pd.isna(time_val) or str(time_val).strip() == "":
        return []

    raw_str = str(time_val).strip()
    meetings = []
    
    # Split by comma if it looks like multiple distinct definitions
    segments = raw_str.split(',')
    
    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue
            
        # 1. Extract Time Range (HH:MM-HH:MM)
        time_match = re.search(r'(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})', segment)
        
        if not time_match:
            continue
            
        start_time = time_match.group(1)
        end_time = time_match.group(2)
        
        # Normalize times to HH:MM
        def normalize_time(t):
            h, m = t.split(':')
            return f"{int(h):02d}:{m}"
        
        start_time = normalize_time(start_time)
        end_time = normalize_time(end_time)
        
        # 2. Remove the time part to process days/venues
        remainder = segment.replace(time_match.group(0), "").strip()
        
        # 3. Strategy: Look for "Day (Venue)" patterns first
        day_venue_pattern = re.compile(r'((?:Th|M|T|W|F|S)+)\s*\(([^)]+)\)')
        matches = list(day_venue_pattern.finditer(remainder))
        matched_spans = []
        
        # Process explicit day-venue pairs
        for m in matches:
            d_str = m.group(1)
            v_str = m.group(2)
            matched_spans.append(m.span())
            
            days = parse_day_string(d_str)
            for d in days:
                meetings.append({
                    'day': d,
                    'start': start_time,
                    'end': end_time,
                    'venue': v_str.strip()
                })
        
        # Process remaining parts (days without explicit venue)
        temp_remainder = list(remainder)
        for start, end in matched_spans:
            for i in range(start, end):
                temp_remainder[i] = ' '
        
        clean_remainder = "".join(temp_remainder).strip()
        
        if clean_remainder:
            days = parse_day_string(clean_remainder)
            venue_to_use = default_venue if pd.notna(default_venue) else "TBA"
            for d in days:
                meetings.append({
                    'day': d,
                    'start': start_time,
                    'end': end_time,
                    'venue': str(venue_to_use).strip()
                })
                
    return meetings

def extract_course_code_title(raw_name):
    """
    Robustly extracts Course Code and Title from various formats.
    """
    raw_name = str(raw_name).strip()
    
    # 1. Try "TITLE (CODE)" or "TITLE(CODE)" at the end (Standard)
    # We look for parens at the end containing something that looks like a code (alphanumeric)
    match = re.search(r'(.*?)\s*\(([^)]+)\)$', raw_name)
    if match:
        title_part = match.group(1).strip()
        code_part = match.group(2).strip()
        
        # Heuristic: Code should be relatively short (e.g. < 12 chars)
        if len(code_part) <= 12:
            return code_part, title_part

    # 2. Try "(SCHEME)(CODE) TITLE" or "TITLE (CODE) SOMETHING"
    # Look for a parenthesized group that contains digits (e.g. ELC113, CS101)
    # This helps distinguish from (SCHEME) or (UG) which have no digits.
    match = re.search(r'\(([A-Z0-9]*\d+[A-Z0-9]*)\)', raw_name)
    if match:
        code = match.group(1).strip()
        # Title is everything else, cleaned up
        title = raw_name.replace(match.group(0), " ").strip()
        # Clean up extra parens or spaces
        title = re.sub(r'\s+', ' ', title)
        # Remove standalone parens like "()" or "( )"
        title = re.sub(r'\(\s*\)', '', title).strip()
        return code, title

    # 3. Try "CODE - TITLE"
    match = re.search(r'^([A-Z0-9]{3,10})\s*-\s*(.*)', raw_name)
    if match:
        code = match.group(1).strip()
        title = match.group(2).strip()
        return code, title

    # Fallback: Use whole string as code (will likely be filtered out by validation later)
    return raw_name, raw_name

def parse_csv_and_normalize(csv_path):
    print(f"Reading CSV from {csv_path}...")
    try:
        df = pd.read_csv(csv_path)
    except FileNotFoundError:
        print(f"Error: File not found at {csv_path}")
        return None, None

    meetings_rows = []
    
    # Identify Time/Venue columns
    time_cols = [c for c in df.columns if str(c).startswith('Time')]
    
    for idx, row in df.iterrows():
        # 1. Basic Course Info
        raw_name = str(row.get('Course Name/Group Name', ''))
        instructor = str(row.get('Instructor', ''))
        
        # Extract Code and Title using improved logic
        course_code, course_title = extract_course_code_title(raw_name)
        
        # --- VALIDATION ---
        # Drop if code is too long or contains invalid characters
        # Valid code: Alphanumeric, maybe hyphens/dots, length 3-10 usually.
        # Let's be generous: length <= 12, no spaces inside code (unless it's like "CS 101")
        
        # Clean code: remove spaces
        course_code = course_code.replace(" ", "")
        
        if len(course_code) > 12 or len(course_code) < 2:
            # Skip this row or log it?
            # We will skip it for the dashboard to be clean.
            continue
            
        # Check for weird characters in code (allow A-Z, 0-9, -, .)
        if not re.match(r'^[A-Z0-9\-\.]+$', course_code, re.IGNORECASE):
            continue
            
        display_course = f"{course_code} - {course_title}"
        
        # 2. Process Time/Venue Pairs
        for t_col in time_cols:
            suffix = t_col.replace('Time', '')
            v_col = 'Venue' + suffix
            
            time_val = row.get(t_col)
            venue_val = row.get(v_col)
            
            if pd.isna(time_val) or str(time_val).strip() == "":
                continue
                
            parsed = parse_time_field(time_val, venue_val)
            
            for m in parsed:
                meetings_rows.append({
                    'CourseCode': course_code,
                    'CourseTitle': course_title,
                    'DisplayCourse': display_course,
                    'Instructor': instructor,
                    'Day': m['day'],
                    'StartTime': m['start'],
                    'EndTime': m['end'],
                    'Venue': m['venue'],
                    'OriginalRawString': raw_name # Added for debugging
                })

    # Create DataFrames
    meetings_df = pd.DataFrame(meetings_rows)
    
    # Create Unique Course List
    course_list_data = []
    seen_codes = set()
    
    for _, row in df.iterrows():
        raw_name = str(row.get('Course Name/Group Name', ''))
        
        course_code, course_title = extract_course_code_title(raw_name)
        course_code = course_code.replace(" ", "")
        
        # Same validation for the course list
        if len(course_code) > 12 or len(course_code) < 2:
            continue
        if not re.match(r'^[A-Z0-9\-\.]+$', course_code, re.IGNORECASE):
            continue
            
        if course_code not in seen_codes:
            course_list_data.append({
                'CourseCode': course_code,
                'CourseTitle': course_title,
                'DisplayCourse': f"{course_code} - {course_title}",
                'Instructor': row.get('Instructor', ''),
                'OriginalRawString': raw_name # Added
            })
            seen_codes.add(course_code)
            
    courses_df = pd.DataFrame(course_list_data)
    if not courses_df.empty:
        courses_df.sort_values('DisplayCourse', inplace=True)
    
    return meetings_df, courses_df

def build_masterdata_sheets(wb, meetings_df, courses_df):
    # 1. MasterCourses Sheet
    if "MasterCourses" in wb.sheetnames:
        ws_courses = wb["MasterCourses"]
    else:
        ws_courses = wb.create_sheet("MasterCourses")
    
    # Write Headers & Data
    ws_courses.append(list(courses_df.columns))
    for r in dataframe_to_rows(courses_df, index=False, header=False):
        ws_courses.append(r)
        
    # Create Table
    tab = Table(displayName="tblCourses", ref=f"A1:{get_column_letter(len(courses_df.columns))}{len(courses_df)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws_courses.add_table(tab)
    
    # 2. MasterMeetings Sheet
    if "MasterMeetings" in wb.sheetnames:
        ws_meetings = wb["MasterMeetings"]
    else:
        ws_meetings = wb.create_sheet("MasterMeetings")
        
    ws_meetings.append(list(meetings_df.columns))
    for r in dataframe_to_rows(meetings_df, index=False, header=False):
        ws_meetings.append(r)
        
    tab_m = Table(displayName="tblMeetings", ref=f"A1:{get_column_letter(len(meetings_df.columns))}{len(meetings_df)+1}")
    tab_m.tableStyleInfo = style
    ws_meetings.add_table(tab_m)

def build_dashboard_layout(ws):
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    
    # Styles
    NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
    TITLE_FONT = Font(color="1F4E78", bold=True, size=18)
    BOLD_FONT = Font(bold=True)
    BORDER = Border(left=Side(style='thin', color="BFBFBF"), 
                    right=Side(style='thin', color="BFBFBF"), 
                    top=Side(style='thin', color="BFBFBF"), 
                    bottom=Side(style='thin', color="BFBFBF"))
    
    # Section A: Title
    ws['A1'] = "📅 IITK Course Timetable (Multi-Course Selector)"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = "Select up to 6 courses. Grid shows 15-minute time slots Mon-Fri, 08:00-19:30"
    
    # Section C: Search Box (Optional)
    ws['H1'] = "Quick Search:"
    ws['H1'].font = BOLD_FONT
    ws['H2'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws['H2'].border = BORDER
    
    # Section B: Course Selection Panel
    ws['A4'] = "Select Courses (Max 6)"
    ws['A4'].font = BOLD_FONT
    ws['E4'] = "Instructor"
    ws['E4'].font = BOLD_FONT
    
    # Dropdowns C5:C10
    for i in range(6):
        row = 5 + i
        cell_c = ws.cell(row=row, column=3) # C
        cell_c.border = BORDER
        cell_c.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        cell_e = ws.cell(row=row, column=5) # E
        cell_e.border = BORDER
        
        # Instructor Lookup Formula
        # =IFERROR(XLOOKUP(LEFT(C5,FIND(" - ",C5)-1), tblCourses[CourseCode], tblCourses[Instructor]), "")
        # Note: We need to handle the case where C5 is empty or invalid format
        formula = f'=IFERROR(XLOOKUP(LEFT(C{row},FIND(" - ",C{row}&" - ")-1), tblCourses[CourseCode], tblCourses[Instructor]), "")'
        cell_e.value = formula

    # Setup Searchable Dropdown Logic
    # We'll use a spilled range in MasterCourses!AA2
    ws_courses = ws.parent["MasterCourses"]
    ws_courses['AA1'] = "Filtered List"
    # Formula: =IFERROR(FILTER(tblCourses[DisplayCourse], ISNUMBER(SEARCH(Dashboard!$H$2, tblCourses[DisplayCourse]))), tblCourses[DisplayCourse])
    ws_courses['AA2'] = '=IFERROR(FILTER(tblCourses[DisplayCourse], ISNUMBER(SEARCH(Dashboard!$H$2, tblCourses[DisplayCourse]))), tblCourses[DisplayCourse])'
    
    # Data Validation
    dv = DataValidation(type="list", formula1='=MasterCourses!$AA$2#', allow_blank=True)
    ws.add_data_validation(dv)
    for i in range(6):
        dv.add(ws.cell(row=5+i, column=3))

    # Section D: Weekly Timetable Grid
    # Headers
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for i, day in enumerate(days):
        cell = ws.cell(row=12, column=2+i)
        cell.value = day
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    # Time Slots
    start_time = datetime.datetime(2000, 1, 1, 8, 0)
    end_time = datetime.datetime(2000, 1, 1, 19, 30)
    current = start_time
    row = 13
    
    while current <= end_time:
        time_str = current.strftime("%H:%M")
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = time_str
        cell_a.font = Font(size=10, bold=True)
        cell_a.alignment = Alignment(horizontal='right')
        cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_a.border = BORDER
        
        # Grid Cells
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        current += datetime.timedelta(minutes=15)
        row += 1
        
    return row # Return the next available row

def inject_grid_formulas(ws, end_row):
    cols = ['B', 'C', 'D', 'E', 'F']
    for col_idx, col_letter in enumerate(cols):
        # Adjust DayName reference for this column
        day_ref = f"{col_letter}$12"
        
        for r in range(13, end_row):
            time_ref = f"$A{r}"
            
            # Reconstruct formula with correct refs
            cell_formula = (
                f'=LET('
                f'SlotStart, TIMEVALUE(LEFT({time_ref}, 5)), '
                f'SlotEnd, SlotStart + TIME(0, 15, 0), '
                f'DayName, {day_ref}, '
                f'C1, IF(ISERROR(FIND(" - ", $C$5)), "", LEFT($C$5, FIND(" - ", $C$5)-1)), '
                f'C2, IF(ISERROR(FIND(" - ", $C$6)), "", LEFT($C$6, FIND(" - ", $C$6)-1)), '
                f'C3, IF(ISERROR(FIND(" - ", $C$7)), "", LEFT($C$7, FIND(" - ", $C$7)-1)), '
                f'C4, IF(ISERROR(FIND(" - ", $C$8)), "", LEFT($C$8, FIND(" - ", $C$8)-1)), '
                f'C5, IF(ISERROR(FIND(" - ", $C$9)), "", LEFT($C$9, FIND(" - ", $C$9)-1)), '
                f'C6, IF(ISERROR(FIND(" - ", $C$10)), "", LEFT($C$10, FIND(" - ", $C$10)-1)), '
                f'SelectedCodes, VSTACK(C1, C2, C3, C4, C5, C6), '
                f'Matching, FILTER('
                f'tblMeetings[CourseCode] & "@" & tblMeetings[Venue], '
                f'(tblMeetings[Day] = DayName) * '
                f'(COUNTIF(SelectedCodes, tblMeetings[CourseCode]) > 0) * '
                f'(TIMEVALUE(LEFT(tblMeetings[StartTime], 5)) < SlotEnd) * '
                f'(TIMEVALUE(LEFT(tblMeetings[EndTime], 5)) > SlotStart), '
                f'""'
                f'), '
                f'IF(AND(ROWS(Matching)=1, INDEX(Matching,1)=""), "", '
                f'IF(ROWS(Matching)=1, INDEX(Matching,1), '
                f'"CONFLICT: " & TEXTJOIN(", ", TRUE, UNIQUE(LEFT(Matching, FIND("@", Matching)-1)))'
                f')))'
            )
            ws[f"{col_letter}{r}"] = cell_formula

def apply_conditional_formatting(ws, end_row):
    grid_range = f"B13:F{end_row-1}"
    
    # Rule 1: Non-empty (Light Blue)
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    ws.conditional_formatting.add(grid_range, 
        FormulaRule(formula=['LEN(B13)>1'], stopIfTrue=False, fill=fill_blue))
        
    # Rule 2: Conflict (Red)
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_red = Font(color="9C0006", bold=True)
    ws.conditional_formatting.add(grid_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("CONFLICT", B13))'], stopIfTrue=True, fill=fill_red, font=font_red))

def add_no_schedule_panel(ws, start_row):
    # Section H: No-Fixed-Schedule Panel
    ws[f'A{start_row}'] = "⚠️ Courses with No Fixed Schedule:"
    ws[f'A{start_row}'].font = Font(bold=True, color="C00000")
    
    # Formula to list courses with no meetings
    formula = (
        f'=TEXTJOIN(", ", TRUE, FILTER($C$5:$C$10, '
        f'(COUNTIF(tblMeetings[CourseCode], LEFT($C$5:$C$10, FIND(" - ", $C$5:$C$10&" - ")-1)) = 0) * '
        f'($C$5:$C$10<>""), ""))'
    )
    
    ws[f'A{start_row+1}'] = formula
    ws.merge_cells(f'A{start_row+1}:F{start_row+2}')
    ws[f'A{start_row+1}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

def finalize_workbook(wb, ws):
    # Column Widths
    ws.column_dimensions['A'].width = 12
    for c in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 20
        
    # Freeze Panes
    ws.freeze_panes = "B13"
    
    # Print Area
    ws.print_area = f"A1:F{ws.max_row}"

def main():
    print("Step 1: Parsing CSV...")
    meetings_df, courses_df = parse_csv_and_normalize(INPUT_CSV)
    
    if meetings_df is None:
        return

    print("Step 2: Creating Workbook...")
    wb = Workbook()
    
    print("Step 3: Building MasterData...")
    build_masterdata_sheets(wb, meetings_df, courses_df)
    
    print("Step 4: Building Dashboard...")
    ws_dash = wb.create_sheet("Dashboard")
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    next_row = build_dashboard_layout(ws_dash)
    
    print("Step 5: Injecting Formulas...")
    inject_grid_formulas(ws_dash, next_row)
    
    print("Step 6: Applying Formatting...")
    apply_conditional_formatting(ws_dash, next_row)
    
    print("Step 7: Adding Extras...")
    add_no_schedule_panel(ws_dash, next_row + 2)
    
    finalize_workbook(wb, ws_dash)
    
    wb.save(OUTPUT_XLSX)
    print(f"✅ Dashboard created: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
