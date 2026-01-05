import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import re
import os

# --- Configuration ---
INPUT_CSV = r"C:\Users\DKS\Desktop\VSC\Dec25\python\Course Schedule 2019-202 Semester(Pre-Registration).csv"
OUTPUT_XLSX = r"C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Dashboard.xlsx"

def parse_course_string(course_str):
    """Extracts Course Code and Name from 'NAME(CODE)' format."""
    if pd.isna(course_str):
        return "", ""
    match = re.search(r'(.*?)\((.*?)\)', str(course_str))
    if match:
        return match.group(2).strip(), match.group(1).strip()
    return "", str(course_str).strip()

def parse_time_venue(time_str, default_venue):
    """
    Parses a time string like 'M (L11) W (L11) F (L11) 09:00-10:00' 
    or 'T (T206) 10:30-11:45 ,Th (T206) 12:00-13:15'
    Returns a list of dictionaries: [{'day': 'Mon', 'time': '09:00-10:00', 'venue': 'L11'}, ...]
    """
    if pd.isna(time_str) or str(time_str).strip() == "":
        return []

    # Split by comma for multiple slots in one cell
    segments = str(time_str).split(',')
    results = []

    for segment in segments:
        segment = segment.strip()
        
        # Extract Time Range
        time_match = re.search(r'(\d{1,2}:\d{2}-\d{1,2}:\d{2})', segment)
        if not time_match:
            continue
        time_slot = time_match.group(1)
        
        # Extract Venue from parens if exists, else use default
        venue_match = re.search(r'\((.*?)\)', segment)
        venue = venue_match.group(1) if venue_match else default_venue
        if pd.isna(venue): venue = ""
        
        # Extract Days
        # Remove time and venue from segment to isolate days
        days_part = segment.replace(time_slot, "")
        if venue_match:
            days_part = days_part.replace(venue_match.group(0), "")
        
        # Normalize Days
        # Replace 'Th' with 'H' to distinguish from 'T'
        days_part = days_part.replace("Th", "H")
        
        days_map = {'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'H': 'Thu', 'F': 'Fri', 'S': 'Sat'}
        found_days = []
        for char in days_part:
            if char in days_map:
                found_days.append(days_map[char])
        
        # If no days found (rare), skip
        if not found_days:
            continue
            
        for day in found_days:
            results.append({
                'day': day,
                'time': time_slot,
                'venue': str(venue).strip()
            })
            
    return results

def main():
    print("Step 1: Reading and Processing Data...")
    try:
        df = pd.read_csv(INPUT_CSV)
    except FileNotFoundError:
        print(f"Error: File not found at {INPUT_CSV}")
        return

    # Prepare Master Data List
    master_data = []

    # Iterate through rows
    for _, row in df.iterrows():
        # Parse Course Info
        raw_course = row.get('Course Name/Group Name', '')
        course_code, course_name = parse_course_string(raw_course)
        instructor = row.get('Instructor', '')
        
        # Process Time/Venue Columns
        # The CSV has pairs: Time, Venue, Time.1, Venue.1, Time.2, Venue.2
        # We'll look for columns starting with 'Time'
        time_cols = [c for c in df.columns if str(c).startswith('Time')]
        
        for t_col in time_cols:
            # Find corresponding venue column (next one usually, or by index)
            # In pandas duplicate names get .1, .2 suffix. 
            # Assuming Venue follows Time immediately in the CSV structure provided.
            # We can try to match suffixes.
            suffix = t_col.replace('Time', '') # '', '.1', '.2'
            v_col = 'Venue' + suffix
            
            time_val = row.get(t_col)
            venue_val = row.get(v_col)
            
            parsed_slots = parse_time_venue(time_val, venue_val)
            
            for slot in parsed_slots:
                master_data.append({
                    'Course Code': course_code,
                    'Course Name': course_name,
                    'Instructor': instructor,
                    'Day': slot['day'],
                    'Time Slot': slot['time'],
                    'Room': slot['venue'],
                    'Full Course': f"{course_code} - {course_name}"
                })

    # Create Master DataFrame
    master_df = pd.DataFrame(master_data)
    master_df.drop_duplicates(inplace=True)
    
    # Standardize Time Slots (Optional: Ensure 08:00 format)
    # The regex already extracts HH:MM-HH:MM.

    print(f"Processed {len(master_df)} schedule entries.")

    # --- Step 2: Excel Creation ---
    print("Step 2: Creating Excel Dashboard...")
    wb = Workbook()
    
    # 1. MasterData Sheet
    ws_master = wb.active
    ws_master.title = "MasterData"
    
    # Write Headers
    headers = list(master_df.columns)
    ws_master.append(headers)
    
    # Write Data
    for r in dataframe_to_rows(master_df, index=False, header=False):
        ws_master.append(r)
        
    # Create Table
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tab = Table(displayName="MasterTable", ref=f"A1:{get_column_letter(len(headers))}{len(master_df)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws_master.add_table(tab)

    # 2. Dashboard Sheet
    ws_dash = wb.create_sheet("Dashboard")
    
    # Styling Constants
    BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws_dash['A1'] = "IITK Course Timetable"
    ws_dash['A1'].font = Font(size=16, bold=True, color="203764")
    
    # Search Box
    ws_dash['A3'] = "Select Course:"
    ws_dash['A3'].font = BOLD_FONT
    ws_dash['C3'].fill = YELLOW_FILL
    ws_dash['C3'].border = BORDER
    
    # Data Validation for C3
    # We need a unique list of courses. 
    # Since XLOOKUP/UNIQUE formulas in validation are tricky across sheets in some excel versions,
    # we'll point to the MasterTable column G (Full Course).
    # A dynamic named range or direct reference to the table column is best.
    # For simplicity, we'll reference the column range in MasterData.
    dv = DataValidation(type="list", formula1=f"=MasterData!$G$2:$G${len(master_df)+1}", allow_blank=True)
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash['C3'])
    
    # Grid Headers
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for i, day in enumerate(days):
        cell = ws_dash.cell(row=5, column=2+i)
        cell.value = day
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    # Grid Rows (Time Slots)
    # We'll generate standard slots or use unique ones from data. 
    # Let's use standard IITK slots roughly.
    time_slots = [
        "08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", 
        "12:00-13:00", "12:00-13:15", "14:00-15:00", "14:00-15:15", 
        "15:00-16:00", "15:30-16:45", "16:00-17:00", "17:00-18:00", "17:15-18:30"
    ]
    # Or better, get unique sorted slots from data to ensure matches
    unique_slots = sorted(master_df['Time Slot'].unique())
    
    for i, slot in enumerate(unique_slots):
        cell = ws_dash.cell(row=6+i, column=1)
        cell.value = slot
        cell.font = BOLD_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='right')

    # Grid Formulas & Formatting
    # Formula: =XLOOKUP(1, (MasterTable[Full Course]=$C$3)*(MasterTable[Day]=B$5)*(MasterTable[Time Slot]=$A6), MasterTable[Room], "")
    # Note: In openpyxl, we write the formula string.
    
    grid_start_row = 6
    grid_end_row = 6 + len(unique_slots) - 1
    grid_start_col = 2 # B
    grid_end_col = 6   # F
    
    for r in range(grid_start_row, grid_end_row + 1):
        for c in range(grid_start_col, grid_end_col + 1):
            cell = ws_dash.cell(row=r, column=c)
            col_letter = get_column_letter(c)
            day_header_ref = f"{col_letter}$5"
            time_row_ref = f"$A{r}"
            
            # Constructing the array formula for XLOOKUP
            # Using MasterTable references
            formula = f'XLOOKUP(1, (MasterTable[Full Course]=$C$3)*(MasterTable[Day]={day_header_ref})*(MasterTable[Time Slot]={time_row_ref}), MasterTable[Room], "")'
            cell.value = f"={formula}"
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Conditional Formatting for the Grid
    # Highlight if not empty
    grid_range = f"B{grid_start_row}:F{grid_end_row}"
    ws_dash.conditional_formatting.add(grid_range, 
        CellIsRule(operator='notEqual', formula=['""'], stopIfTrue=True, fill=LIGHT_BLUE_FILL))

    # Info Panel
    info_row = grid_end_row + 3
    ws_dash[f'A{info_row}'] = "Instructor:"
    ws_dash[f'A{info_row}'].font = BOLD_FONT
    ws_dash[f'C{info_row}'] = f'=XLOOKUP($C$3, MasterTable[Full Course], MasterTable[Instructor], "Not Found")'
    
    ws_dash[f'A{info_row+1}'] = "Course Name:"
    ws_dash[f'A{info_row+1}'].font = BOLD_FONT
    ws_dash[f'C{info_row+1}'] = f'=XLOOKUP($C$3, MasterTable[Full Course], MasterTable[Course Name], "Not Found")'

    # Final Styling
    ws_dash.column_dimensions['A'].width = 18
    for c in ['B', 'C', 'D', 'E', 'F']:
        ws_dash.column_dimensions[c].width = 20
    
    ws_dash.sheet_view.showGridLines = False

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"Success! Dashboard saved to: {OUTPUT_XLSX}")

def dataframe_to_rows(df, index=True, header=True):
    """
    Simple utility to yield rows from dataframe for openpyxl
    """
    import numpy as np
    if header:
        yield [index_name for index_name in df.index.names] + list(df.columns) if index else list(df.columns)
    for row in df.itertuples(index=index, name=None):
        yield [None if pd.isna(x) else x for x in row]

if __name__ == "__main__":
    main()
