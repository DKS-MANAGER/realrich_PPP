import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import re
import os

# --- Configuration ---
INPUT_CSV = r"C:\Users\DKS\Desktop\VSC\Dec25\python\Course Schedule 2019-202 Semester(Pre-Registration).csv"
OUTPUT_XLSX = r"C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Upgraded.xlsx"

def parse_course_string(course_str):
    """Extracts Course Code and Name from 'NAME(CODE)' format."""
    if pd.isna(course_str):
        return "", ""
    match = re.search(r'(.*?)\((.*?)\)', str(course_str))
    if match:
        return match.group(2).strip(), match.group(1).strip()
    return "", str(course_str).strip()

def parse_time_venue(time_str, default_venue):
    """
    Parses a time string like 'M (L11) W (L11) F (L11) 09:00-10:00' 
    or 'T (T206) 10:30-11:45 ,Th (T206) 12:00-13:15'
    Returns a list of dictionaries: [{'day': 'Mon', 'time': '09:00-10:00', 'venue': 'L11'}, ...]
    """
    if pd.isna(time_str) or str(time_str).strip() == "":
        return []

    # Split by comma for multiple slots in one cell
    segments = str(time_str).split(',')
    results = []

    for segment in segments:
        segment = segment.strip()
        
        # Extract Time Range
        time_match = re.search(r'(\d{1,2}:\d{2}-\d{1,2}:\d{2})', segment)
        if not time_match:
            continue
        time_slot = time_match.group(1)
        
        # Extract Venue from parens if exists, else use default
        venue_match = re.search(r'\((.*?)\)', segment)
        venue = venue_match.group(1) if venue_match else default_venue
        if pd.isna(venue): venue = ""
        
        # Extract Days
        # Remove time and venue from segment to isolate days
        days_part = segment.replace(time_slot, "")
        if venue_match:
            days_part = days_part.replace(venue_match.group(0), "")
        
        # Normalize Days
        # Replace 'Th' with 'H' to distinguish from 'T'
        days_part = days_part.replace("Th", "H")
        
        days_map = {'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'H': 'Thu', 'F': 'Fri', 'S': 'Sat'}
        found_days = []
        for char in days_part:
            if char in days_map:
                found_days.append(days_map[char])
        
        # If no days found (rare), skip
        if not found_days:
            continue
            
        for day in found_days:
            results.append({
                'day': day,
                'time': time_slot,
                'venue': str(venue).strip()
            })
            
    return results

def main():
    print("Step 1: Reading and Processing Data...")
    try:
        df = pd.read_csv(INPUT_CSV)
    except FileNotFoundError:
        print(f"Error: File not found at {INPUT_CSV}")
        return

    # Prepare Master Data List
    master_data = []

    # Iterate through rows
    for _, row in df.iterrows():
        # Parse Course Info
        raw_course = row.get('Course Name/Group Name', '')
        course_code, course_name = parse_course_string(raw_course)
        instructor = row.get('Instructor', '')
        
        # Process Time/Venue Columns
        time_cols = [c for c in df.columns if str(c).startswith('Time')]
        
        for t_col in time_cols:
            suffix = t_col.replace('Time', '') # '', '.1', '.2'
            v_col = 'Venue' + suffix
            
            time_val = row.get(t_col)
            venue_val = row.get(v_col)
            
            parsed_slots = parse_time_venue(time_val, venue_val)
            
            for slot in parsed_slots:
                master_data.append({
                    'Course Code': course_code,
                    'Course Name': course_name,
                    'Instructor': instructor,
                    'Day': slot['day'],
                    'Time Slot': slot['time'],
                    'Room': slot['venue'],
                    'Full Course': f"{course_code} - {course_name}"
                })

    # Create Master DataFrame
    master_df = pd.DataFrame(master_data)
    master_df.drop_duplicates(inplace=True)
    
    print(f"Processed {len(master_df)} schedule entries.")

    # --- Step 2: Excel Creation ---
    print("Step 2: Creating Upgraded Excel Dashboard...")
    wb = Workbook()
    
    # 1. MasterData Sheet
    ws_master = wb.active
    ws_master.title = "MasterData"
    
    # Write Headers
    headers = list(master_df.columns)
    ws_master.append(headers)
    
    # Write Data
    for r in dataframe_to_rows(master_df, index=False, header=False):
        ws_master.append(r)
        
    # Create Table
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tab = Table(displayName="MasterTable", ref=f"A1:{get_column_letter(len(headers))}{len(master_df)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws_master.add_table(tab)

    # 2. Dashboard Sheet
    ws_dash = wb.create_sheet("Dashboard")
    
    # Styling Constants
    NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    LIGHT_GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
    TITLE_FONT = Font(size=24, bold=True, color="1F4E78")
    BOLD_FONT = Font(bold=True)
    
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- LAYOUT BUILD ---
    # Title
    ws_dash["B2"] = "📅 IITK Course Timetable"
    ws_dash["B2"].font = TITLE_FONT

    # Search Box
    ws_dash["B4"] = "Select Course:"
    ws_dash["B4"].font = Font(bold=True, size=12)
    ws_dash["C4"].font = Font(size=12)
    ws_dash["C4"].border = THIN_BORDER
    ws_dash["C4"].fill = YELLOW_FILL

    # Data Validation
    # Pointing to Column G (Full Course) in MasterData
    # Using a formula reference to the entire column G in MasterData table would be ideal, 
    # but for compatibility we use the range reference.
    dv = DataValidation(type="list", formula1=f"=MasterData!$G$2:$G${len(master_df)+1}", allow_blank=True)
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash["C4"])

    # Grid Headers
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for i, day in enumerate(days):
        col = i + 3 # Start at Column C (3)
        cell = ws_dash.cell(row=6, column=col)
        cell.value = day
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Time Slots (Rows)
    times = ["08:00", "09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00", "17:00", "18:00"]
    for i, time in enumerate(times):
        row = i + 7
        cell = ws_dash.cell(row=row, column=2) # Column B
        cell.value = time
        cell.font = BOLD_FONT
        cell.fill = LIGHT_GREY_FILL
        cell.alignment = RIGHT_ALIGN
        cell.border = THIN_BORDER

    # --- FORMULAS ---
    # Grid Formula
    # We use LEFT($B7, 2) to match the hour part of the time slot in MasterData
    # MasterData Columns: A=Code, D=Day, E=Time, F=Room
    
    for row_idx in range(7, 7 + len(times)):
        for col_idx in range(3, 8): # C to G (Mon to Fri)
            cell = ws_dash.cell(row=row_idx, column=col_idx)
            
            col_letter = get_column_letter(col_idx)
            day_header_ref = f"{col_letter}$6"
            time_row_ref = f"$B{row_idx}"
            
            # Formula Logic:
            # 1. Extract Course Code from C4: LEFT($C$4, FIND(" - ", $C$4&" - ")-1)
            # 2. Filter MasterData based on:
            #    - Course Code matches
            #    - Day matches column header
            #    - Time matches row label (using SEARCH on the hour)
            # 3. TEXTJOIN the results (Rooms)
            
            formula = (
                f'=IF($C$4="","",IFERROR(TEXTJOIN(", ",TRUE,FILTER(MasterData!$F:$F,'
                f'(MasterData!$A:$A=LEFT($C$4,FIND(" - ",$C$4&" - ")-1))*'
                f'(MasterData!$D:$D={day_header_ref})*'
                f'(ISNUMBER(SEARCH(LEFT({time_row_ref},2),MasterData!$E:$E))))),""))'
            )
            
            cell.value = formula
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

    # --- CONDITIONAL FORMATTING ---
    # Highlight non-empty cells in the grid
    grid_range = f"C7:G{7+len(times)-1}"
    ws_dash.conditional_formatting.add(grid_range, 
        CellIsRule(operator="notEqual", formula=['""'], stopIfTrue=True, fill=LIGHT_BLUE_FILL))

    # --- INFO PANEL ---
    info_row = 7 + len(times) + 2
    ws_dash[f'B{info_row}'] = "Instructor:"
    ws_dash[f'B{info_row}'].font = BOLD_FONT
    ws_dash[f'C{info_row}'] = f'=IF($C$4="","",XLOOKUP(LEFT($C$4,FIND(" - ",$C$4&" - ")-1), MasterData!A:A, MasterData!C:C, "Not Found"))'
    
    ws_dash[f'B{info_row+1}'] = "Course Name:"
    ws_dash[f'B{info_row+1}'].font = BOLD_FONT
    ws_dash[f'C{info_row+1}'] = f'=IF($C$4="","",XLOOKUP(LEFT($C$4,FIND(" - ",$C$4&" - ")-1), MasterData!A:A, MasterData!B:B, "Not Found"))'

    # --- FINAL STYLING ---
    ws_dash.column_dimensions['A'].width = 2
    ws_dash.column_dimensions['B'].width = 15
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws_dash.column_dimensions[c].width = 20
    
    ws_dash.sheet_view.showGridLines = False

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"Success! Upgraded Dashboard saved to: {OUTPUT_XLSX}")

def dataframe_to_rows(df, index=True, header=True):
    import numpy as np
    if header:
        yield [index_name for index_name in df.index.names] + list(df.columns) if index else list(df.columns)
    for row in df.itertuples(index=index, name=None):
        yield [None if pd.isna(x) else x for x in row]

if __name__ == "__main__":
    main()
