import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import os

# Config
FILE_PATH = r'C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Dashboard_v3.xlsx'
OUTPUT_PATH = r'C:\Users\DKS\Desktop\VSC\Dec25\python\IITK_Timetable_Dashboard_v3_FIXED.xlsx'

# Colors
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Dark Blue
WHITE_FONT = Font(color='FFFFFF', bold=True)
TIME_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') # Light Gray
BORDER = Border(left=Side(style='thin', color='BFBFBF'), 
                right=Side(style='thin', color='BFBFBF'), 
                top=Side(style='thin', color='BFBFBF'), 
                bottom=Side(style='thin', color='BFBFBF'))

def repair_dashboard():
    print(f"Loading {FILE_PATH}...")
    if not os.path.exists(FILE_PATH):
        print(f"Error: File not found at {FILE_PATH}")
        return

    wb = openpyxl.load_workbook(FILE_PATH)
    
    # 1. Reset Dashboard Sheet
    if 'Dashboard' in wb.sheetnames:
        # Clear existing cells to be safe
        ws = wb['Dashboard']
        wb.remove(ws)
    ws = wb.create_sheet('Dashboard', 0) # Create fresh at index 0

    print("Rebuilding UI...")

    # 2. Fix Layout & Column Widths (CRITICAL FIX)
    ws.column_dimensions['A'].width = 12  # Time column
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20  # Day columns (Fixed "Squished" bug)

    # 3. Title & Header
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "📅 IITK Course Timetable (Multi-Course Selector)"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = "Select up to 6 courses below. Grid shows 15-minute intervals. Red cells indicate conflicts."
    ws.merge_cells('A2:F2')
    ws['A2'].font = Font(italic=True, size=10)
    
    # 4. Course Selector Area
    ws['C4'] = "Select Courses (Type to Search)"
    ws['C4'].font = Font(bold=True)
    ws['E4'] = "Instructor"
    ws['E4'].font = Font(bold=True)

    # Add Data Validation (Dropdowns)
    # Note: INDIRECT might trigger warnings in some Excel versions if table isn't loaded, but usually fine.
    dv = DataValidation(type="list", formula1="=INDIRECT(\"tblCourses[DisplayCourse]\")", allow_blank=True)
    ws.add_data_validation(dv)

    for i in range(5, 11): # Rows 5-10
        # Input Cell
        cell = ws[f'C{i}']
        cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid') # Light Yellow
        cell.border = BORDER
        dv.add(cell)
        
        # Instructor Lookup Formula
        # Using VLOOKUP on tblCourses. tblCourses has DisplayCourse at col 3, Instructor at col 4.
        # Wait, tblCourses columns: CourseCode, CourseTitle, DisplayCourse, Instructor.
        # DisplayCourse is col 3. Instructor is col 4.
        # VLOOKUP(lookup_value, table_array, col_index, [range_lookup])
        # If we look up DisplayCourse in tblCourses, we need DisplayCourse to be the first column of the range?
        # tblCourses range is A:D. DisplayCourse is C.
        # VLOOKUP won't work if we look up value in C and want value in D, unless we select range C:D.
        # Let's use XLOOKUP or INDEX/MATCH for robustness if possible, or adjust VLOOKUP.
        # The user provided script uses: VLOOKUP(LEFT(C{i}, FIND(" - ",C{i}&" - ")-1), tblCourses, 4, FALSE)
        # LEFT(...) extracts CourseCode.
        # CourseCode is column 1 (A). Instructor is column 4 (D).
        # So VLOOKUP(Code, A:D, 4, FALSE) works perfectly.
        
        ws[f'E{i}'] = f'=IFERROR(VLOOKUP(LEFT(C{i}, FIND(" - ",C{i}&" - ")-1), tblCourses, 4, FALSE), "")'
        ws[f'E{i}'].font = Font(italic=True, color='555555')

    # 5. Grid Headers
    headers = ['Time', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=12, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center')

    # 6. Inject Grid Formulas (CRITICAL FIX)
    # Time slots from 08:00 to 19:30 (48 slots)
    start_row = 13
    t = datetime.datetime(2000, 1, 1, 8, 0) # 08:00
    
    # The Complex LET Formula
    # Note: We need to be careful with newlines in f-strings or just string concatenation.
    grid_formula = (
        '=IFERROR(LET('
        'SlotStart, TIMEVALUE($A{row}), '
        'SlotEnd, SlotStart + TIME(0,15,0), '
        'DayName, {day_col}$12, '
        'Codes, FILTER(LEFT($C$5:$C$10, IFERROR(FIND(" - ",$C$5:$C$10)-1,LEN($C$5:$C$10))), $C$5:$C$10<>""), '
        'Match, FILTER(tblMeetings[DisplayCourse] & " (" & tblMeetings[Venue] & ")", '
        '(tblMeetings[Day]=DayName) * '
        '(COUNTIF(Codes, tblMeetings[CourseCode])) * '
        '(TIMEVALUE(tblMeetings[StartTime]) < SlotEnd) * '
        '(TIMEVALUE(tblMeetings[EndTime]) > SlotStart), '
        '""'
        '), '
        'IF(ROWS(Match)=0, "", IF(ROWS(Match)=1, INDEX(Match,1), "CONFLICT!"))'
        '), "")'
    )
    
    print("Injecting 400+ formulas...")
    for i in range(48): # 48 slots * 15 min = 12 hours
        r = start_row + i
        time_str = t.strftime("%H:%M")
        
        # Time Label
        ws.cell(row=r, column=1, value=time_str).fill = TIME_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
        
        # Grid Cells
        for col_idx, day_col in enumerate(['B', 'C', 'D', 'E', 'F'], 2):
            cell = ws.cell(row=r, column=col_idx)
            # Clean up formula string
            f = grid_formula.replace('{row}', str(r)).replace('{day_col}', day_col)
            cell.value = f
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        t += datetime.timedelta(minutes=15)

    # 7. Final Polish
    ws.freeze_panes = 'B13'
    ws.sheet_view.showGridLines = False
    
    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print("✅ Done! Open the FIXED file.")

if __name__ == "__main__":
    repair_dashboard()
